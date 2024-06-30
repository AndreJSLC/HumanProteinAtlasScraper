# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 15:29:22 2024

@author: Maria Carneiro
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import time
import pandas as pd
from openpyxl.styles import PatternFill
# Initialize the webdriver
options = Options()
options.add_argument("-profile")
options.add_argument(r"C:\Users\Maria Carneiro\AppData\Roaming\Mozilla\Firefox\Profiles\steamscraperAlpha")
driver = webdriver.Firefox(options=options)

gene_urls = [
    "https://www.proteinatlas.org/ENSG00000121966-CXCR4/cell+line",
    "https://www.proteinatlas.org/ENSG00000160791-CCR5/cell+line",
    "https://www.proteinatlas.org/ENSG00000088305-DNMT3B/cell+line",
    "https://www.proteinatlas.org/ENSG00000112715-VEGFA/cell+line",
    "https://www.proteinatlas.org/ENSG00000244734-HBB/cell+line",
    "https://www.proteinatlas.org/ENSG00000211772-TRBC2/cell+line",
    "https://www.proteinatlas.org/ENSG00000119866-BCL11A/cell+line",
    "https://www.proteinatlas.org/ENSG00000197386-HTT/cell+line",
    "https://www.proteinatlas.org/ENSG00000198707-CEP290/cell+line",
    "https://www.proteinatlas.org/ENSG00000105610-KLF1/cell+line",
    "https://www.proteinatlas.org/ENSG00000171241-SHCBP1/cell+line",
    "https://www.proteinatlas.org/ENSG00000171431-KRT20/cell+line",
    "https://www.proteinatlas.org/ENSG00000139292-LGR5/cell+line",
    "https://www.proteinatlas.org/ENSG00000135638-EMX1/cell+line",
    "https://www.proteinatlas.org/ENSG00000277734-TRAC/cell+line",
    "https://www.proteinatlas.org/ENSG00000130638-ATXN10/cell+line",
    "https://www.proteinatlas.org/ENSG00000157227-MMP14/cell+line",
    "https://www.proteinatlas.org/ENSG00000170017-ALCAM/cell+line",
    "https://www.proteinatlas.org/ENSG00000166710-B2M/cell+line",
    "https://www.proteinatlas.org/ENSG00000206503-HLA-A/cell+line",
    "https://www.proteinatlas.org/ENSG00000169174-PCSK9/cell+line",
    "https://www.proteinatlas.org/ENSG00000188389-PDCD1/cell+line",
    "https://www.proteinatlas.org/ENSG00000164985-PSIP1/cell+line",
    "https://proteinatlas.org/ENSG00000128294-TPST2/cell+line",
    "https://www.proteinatlas.org/ENSG00000211751-TRBC1/cell+line",
    "https://www.proteinatlas.org/ENSG00000157593-SLC35B2/cell+line",
    "https://www.proteinatlas.org/ENSG00000163631-ALB/cell+line",
    "https://www.proteinatlas.org/ENSG00000001626-CFTR/cell+line",
    "https://www.proteinatlas.org/ENSG00000114270-COL7A1/cell+line",
    "https://www.proteinatlas.org/ENSG00000165168-CYBB/cell+line",
    "https://www.proteinatlas.org/ENSG00000198947-DMD/cell+line",
    "https://www.proteinatlas.org/ENSG00000148935-GAS2/cell+line",
    "https://www.proteinatlas.org/ENSG00000147168-IL2RG/cell+line",
    "https://www.proteinatlas.org/ENSG00000197249-SERPINA1/cell+line",
    "https://www.proteinatlas.org/ENSG00000159216-RUNX1/cell+line",
    "https://www.proteinatlas.org/ENSG00000163599-CTLA4/cell+line",
    "https://www.proteinatlas.org/ENSG00000119888-EPCAM/cell+line",
    "https://www.proteinatlas.org/ENSG00000075891-PAX2/cell+line",
    "https://www.proteinatlas.org/ENSG00000130592-LSP1/cell+line",
    "https://www.proteinatlas.org/ENSG00000007062-PROM1/cell+line",
    "https://www.proteinatlas.org/ENSG00000171431-KRT20/cell+line",
    "https://www.proteinatlas.org/ENSG00000211772-TRBC2/cell+line",
    "https://www.proteinatlas.org/ENSG00000026103-FAS/cell+line",
    "https://www.proteinatlas.org/ENSG00000007062-PROM1/cell+line",
    "https://www.proteinatlas.org/ENSG00000171431-KRT20/cell+line"
    
    
]

#remove duplicates in case they existe while keeping the order
gene_urls = list(dict.fromkeys(gene_urls))

max_retries = 3
data_frames = {}

for gene_url in gene_urls:
    retry_count = 0
    gene_name = gene_url.split("/")[-2]

    while retry_count < max_retries:
        try:
            driver.get(gene_url)
            time.sleep(5)

            if "Service Unavailable" in driver.page_source:
                print("Service Unavailable error detected at", gene_url, ". Retrying...")
                retry_count += 1
                time.sleep(5)
                continue

            # Scrape gene expression data
            cell_lines = ["U2OS", "HeLa", "HTERT-RPE1", "MCF-7", "HPB-ALL", "Jurkat E6.1", "JURKAT", "Huh-7", "K-562", "A-549", "MDA-MB-231", "HaCaT", "THP-1", "DND-41", "HEK293", "G-292 clone A141B1", "NCI-H2444", "SUM159PT", "Ramos", "OVCAR-8","H4","TMD8","OCI-Ly3","HOS","SaOS-2"]                                                          
            expression_status = {}

            svg_elements = driver.find_elements(By.CSS_SELECTOR, "svg.barchart g.bar_g")
            for element in svg_elements:
                title_text = element.get_attribute('title')
                if title_text:
                    for cell_line in cell_lines:
                        if cell_line in title_text:
                            match = re.search(r'nTPM: ([\d\.]+)', title_text)
                            if match:
                                nTPM_value = float(match.group(1))
                                expression_status[cell_line] = ('+' if nTPM_value > 0 else '-', nTPM_value)
                            else:
                                expression_status[cell_line] = ('Data not found', 0)

            df = pd.DataFrame.from_dict(expression_status, orient='index', columns=['Status', 'nTPM'])
            df = df.reset_index().rename(columns={'index': 'Cell Line'})
            df.sort_values(by='nTPM', ascending=False, inplace=True)

            # Scrape cell line distribution data
            cell_line_distribution = None
            try:
                distribution_css_selector = "body > table > tbody > tr > td:nth-child(2) > div > table:nth-child(2) > tbody > tr > td:nth-child(2) > table > tbody > tr:nth-child(5) > td"
                distribution_element = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, distribution_css_selector))
                )
                cell_line_distribution = distribution_element.text
            except Exception as e:
                print(f"Could not find cell line distribution for {gene_name}: {e}")

            data_frames[gene_name] = (df, cell_line_distribution)
            break

        except Exception as e:
            print(f"An error occurred at {gene_url}: {e}")
            retry_count += 1
            time.sleep(10)

    if retry_count == max_retries:
        print("Maximum retries reached at", gene_url, ". Exiting.")

driver.quit()

# Create an Excel writer using openpyxl engine within a context manager
excel_file_path = 'C:\\ProteinAtlasScraperResults\\Gene_Expression_Data.xlsx'
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as excel_writer:
    for gene_name, (df, cell_line_distribution) in data_frames.items():
        # Start writing the DataFrame from row 3 to leave space for the cell line distribution
        start_row = 3
        df.to_excel(excel_writer, sheet_name=gene_name, startrow=start_row, index=False)

        worksheet = excel_writer.sheets[gene_name]

        # Insert cell line distribution data above the table if available
        if cell_line_distribution:
            worksheet.cell(row=1, column=1).value = f"Cell Line Distribution for {gene_name}:"
            worksheet.cell(row=2, column=1).value = cell_line_distribution

        # Apply color formatting
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in worksheet.iter_rows(min_row=start_row+1, max_col=3, max_row=worksheet.max_row):
            if row[1].value == '+':
                for cell in row:
                    cell.fill = green_fill
            elif row[1].value == '-':
                for cell in row:
                    cell.fill = red_fill
